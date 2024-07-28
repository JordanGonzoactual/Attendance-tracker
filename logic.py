import smtplib
import openpyxl
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os 
import logging

# loads environmental variables
email_user = os.getenv('EMAIL_USER')
email_pass = os.getenv('EMAIL_PASS')
print(f"EMAIL_USER: {email_user}")
print(f"EMAIL_PASS: {email_pass}")
#verifies environmental variables are set
if not email_user or not email_pass:
    raise EnvironmentError("EMAIL_USER AND EMAIL_PASS environment variables must be set")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class Student:
    def __init__(self, name, email, status, days):
        self.name = name
        self.email = email
        self.status = status 
        self.no_of_days = days if days else 0
       # Marks students late 
    def mark_late(self):
        self.status = 'Late'
        self.no_of_days +=1

# To send email
    def send_email(self, subject, body):
        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = self.email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        try:
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()# start TLS for security
                server.login(os.getenv('EMAIL_USER'), os.getenv('EMAIL_PASS'), msg.as_string()) #logs in to staff email account
                server.sendmail(self.email, email_user, msg)
                logging.info(f"Email sent to {self.name} at {self.email}")
                print(f"Email sent to {self.name} at {self.email}")
        except smtplib.SMTPAuthenticationError as e:
            logging.error(f"SMTP Authentication Error: Please check your email and password or App password settings")
        except Exception as e:
            logging.error("Failed to send email to {self.name} at {self.email}: {e}")
    
    # for students if missed days
    def notify_late(self, *args):
        if self.status == 'Late':
            subject = "Attendance warning"
            body = (f"Dear {self.name},\n\n"
                f"You have missed {self.no_of_days} days of class. " 
                "Please be aware that you are approaching the max days allowed to miss."
                "Best Regards, \n"
                "Attendance office")
            self.send_email(subject, body)
            


    def __str__(self):
        return f'Student(Name: {self.name}, Email: {self.email}, Status: {self.status}, No of Days Late: {self.no_of_days})'

def validate_excel_structure(sheet, expected_columns):
    header= [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    missing_columns = [col for col in expected_columns if col not in header]
    if missing_columns:
        raise ValueError(f"Missing expected columns: {','.join(missing_columns)}")
    logging.info("Excel structure validated succesfully.")


# for processing attendance
def process_attendance(sheet):
    logging.info("Starting to process attendance")
    students =[]
# iterates over the rows in excel
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4, values_only=True):
        name, email, status, no_of_days = row
        logging.info(f"Processing student: {name}, {email}, {status}, {no_of_days}")
        student = Student(name, email, status, no_of_days)
        
        if student.status.lower() == 'late':
            student.mark_late()
            student.notify_late(name, student, email)
        students.append(student)
    return students


# saves excel on every update
def save_workbook(workbook, filename):
    try:
        workbook.save(filename)
        print("saved!")
    except Exception as e:
        logging.error(f'Error saving workbook: {e}') 
# staff emails
staff_mail=['targyarg13@gmail.com']
# Max amount of missed days
attendance_threshold= 3
# number of days students have missed
status = no_of_days= []
# list of students to remind
l1 =[]
# warning messages
m1 = "Warning!!! you can only miss more day for CI class"
m2 = " Warning !!! you can only miss one more day for python class"
m3 = "Warning!!! you can only miss one more day for DM class"

# main script
if __name__ == "__main__":
    
    try:
        file_path = r'H:\Python\Attendancetracker\attendance.xlsx'
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} does not exist.")
        

        workbook = load_workbook(file_path)
        logging.info(f"Available worksheets: {workbook.sheetnames}")
        
        if "attendance" not in workbook.sheetnames:
            raise ValueError("Worksheet 'attendance' does not exist.")
         
        sheet = workbook["attendance"] 
        #validates structure of excel file
        expected_columns = ["Name", "Email", "Status", "No_of_Days"]
        validate_excel_structure(sheet, expected_columns)
       
        process_attendance(sheet)
        save_workbook(workbook, file_path)
    except Exception as e:
        logging.error(f"Failed to run the sript: {e}")
    finally:
        print("Script finished successfully.")
       

